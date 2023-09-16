import os
import subprocess
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

base_dir = Path(__file__).parent
excel_path = base_dir / "chat_log.xlsx"


def is_output_open_excel() -> bool:
    """
    Excelファイルが開かれているかどうかを判定する
    :return:
    """

    # Windows
    if os.name == "nt":
        try:
            with excel_path.open("r+b"):
                return False
        except PermissionError:
            return True

    # Mac
    if os.name == "posix":
        if excel_path.exists():
            result = subprocess.run(["lsof", excel_path], stdout=subprocess.PIPE)
            return bool(result.stdout)
        return False


def load_or_create_workbook() -> tuple[openpyxl.Workbook, bool]:
    """
    Excelファイルを読み込むか、存在しない場合は作成する
    :returns: ワークブックオブジェクトと、ファイルが作成されたかどうかのフラグ
    """

    # ファイルの存在確認
    if excel_path.exists():
        # ファイルを読み込んで返す
        wb = openpyxl.load_workbook(excel_path)
        return wb, False
    else:
        # ファイルを作成して返す
        wb = openpyxl.Workbook()
        return wb, True


def create_worksheet(title: str, target_workbook: openpyxl.Workbook, is_new: bool):
    """
    シートを作成してシート名に使えない文字を除去したうえでシート名を変更して返す
    :param title: シート名(プロンプトの要約)
    :param target_workbook: 対象になるワークブックオブジェクト
    :param is_new: ワークブックが新規作成されたかどうかのフラグ
    :return: ワークシートオブジェクト
    """

    # シート名に使えない文字を除去
    trimmed_title = trim_invalid_chars(title)

    if is_new:
        # アクティブシート(Sheet)を取得
        target_worksheet = target_workbook.active
        target_worksheet.title = trimmed_title
    else:
        # シートを追加
        target_worksheet = target_workbook.create_sheet(title=trimmed_title)
        target_workbook.move_sheet(target_worksheet, offset=-len(target_workbook.worksheets) + 1)
        target_workbook.active = target_worksheet

    return target_worksheet


def trim_invalid_chars(title: str) -> str:
    """
    シート名に使えない文字を除去する
    :param title: シート名
    :return: 除去後のシート名
    """

    new_title = title
    invalid_chars = ["/", "\\", "?", "*", "[", "]"]
    for char in invalid_chars:
        new_title = new_title.replace(char, "")
    return new_title


def header_formatting(target_worksheet):
    """
    出力対象のワークシートにヘッダーを設定する
    :param target_worksheet: 出力対象のワークシート
    """

    # A1セルに出力時の日時を書き込みフォントを設定
    datetime_cell = target_worksheet["A1"]
    datetime_cell.value = datetime.now().strftime("%Y/%m/%d %H:%M")
    datetime_cell.font = Font(name="メイリオ")

    # A2セルに「ロール」B2セルに「発言内容」と書き込み、フォントとセルの色を設定
    role_header_cell = target_worksheet["A2"]
    content_header_cell = target_worksheet["B2"]

    # セルに値を設定
    role_header_cell.value = "ロール"
    content_header_cell.value = "発言内容"

    # フォントの設定
    white_color = "FFFFFF"
    header_font_style = Font(name="メイリオ", bold=True, color=white_color)
    role_header_cell.font = header_font_style
    content_header_cell.font = header_font_style

    # セルの色を設定
    excel_green = "156B31"
    header_color = PatternFill(fill_type="solid", fgColor=excel_green)
    role_header_cell.fill = header_color
    content_header_cell.fill = header_color

    # セルの幅を調整
    target_worksheet.column_dimensions["A"].width = 22
    target_worksheet.column_dimensions["B"].width = 168


def write_chat_log(target_worksheet, chat_log: list[dict]):
    """
    チャットの履歴を書き込み書式設定する
    :param target_worksheet: 出力対象のワークシート
    :param chat_log: チャットの履歴
    """

    row_height_adjustment_standard = 17
    font_style = Font(name="メイリオ", size=10)
    light_gray = "d9d9d9"
    assistant_color = PatternFill(fill_type="solid", fgColor=light_gray)

    # チャット内容の書き込み
    write_start_row = 3
    for row_number, content in enumerate(chat_log, write_start_row):
        cell_role, cell_content = target_worksheet[f"A{row_number}"], target_worksheet[f"B{row_number}"]

        # ロールと発言内容を書き込み
        cell_role.value = content["role"]
        cell_content.value = content["content"]

        # セル内改行の調整
        cell_content.alignment = Alignment(wrapText=True)

        # 行の高さを調整
        adjusted_row_height = len(content["content"].split("\n")) * row_height_adjustment_standard
        target_worksheet.row_dimensions[row_number].height = adjusted_row_height

        # 書式設定
        cell_role.font = font_style
        cell_content.font = font_style
        if content["role"] == "assistant":
            cell_role.fill = assistant_color
            cell_content.fill = assistant_color


def open_workbook():
    """Excelファイルを開く"""

    # Windows
    if os.name == "nt":
        os.system(f"start {excel_path}")

    # Mac
    if os.name == "posix":
        os.system(f"open {excel_path}")


def output_excel(chat_log: list[dict], chat_summary: str):
    """
    chat_history.xlsx にチャットの履歴を書き込むためのエントリポイント。
    :param chat_log: チャットの履歴
    :param chat_summary: チャットの要約
    :return:
    """
    workbook, is_created = load_or_create_workbook()
    worksheet = create_worksheet(title=chat_summary, target_workbook=workbook, is_new=is_created)
    header_formatting(target_worksheet=worksheet)
    write_chat_log(target_worksheet=worksheet, chat_log=chat_log)
    workbook.save(excel_path)
    workbook.close()
    open_workbook()


if __name__ == "__main__":
    log = [
        {"role": "user", "content": "こんにちは"},
        {"role": "assistant", "content": "こんにちは"},
        {"role": "user", "content": "元気ですか？"},
        {"role": "assistant", "content":
            "1. 元気\nです\nあり\nがと\nう\nござい\nます\nおかげ\nさまで\n元気です\nあなたは\n元気\nですか？"
            "\n2. 元気\nです\nあり\nがと\nう\nござい\nます\nおかげ\nさまで\n元気です\nあなたは\n元気\nですか？"
            "\n3. 元気\nです\nあり\nがと\nう\nござい\nます\nおかげ\nさまで\n元気です\nあなたは\n元気\nですか？"},
    ]

    output_excel(log, "test/\\?*[]")
